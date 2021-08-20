import numpy as np
import pandas as pd
from tqdm.notebook import tqdm
import matplotlib.pyplot as plt
import seaborn as sns

import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader, WeightedRandomSampler

from sklearn.preprocessing import MinMaxScaler, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix, classification_report
from sklearn.model_selection import KFold
import openpyxl
wb = openpyxl.Workbook()
import datetime
import copy
import os

# Time for saving data
currentDT = datetime.datetime.now()
# Load data
df = pd.read_excel(r"C:\Users\admin\Dropbox\Research\개인연구\23_Brachy_BladderTocixity\GU toxicity_DB_python-balance-2-1.xlsx",engine='openpyxl') # Best performance data set
# df = pd.read_excel(r"C:\Users\admin\Dropbox\Research\개인연구\23_Brachy_BladderTocixity\GU toxicity_DB_python-balance.xlsx",engine='openpyxl')
# df = pd.read_excel(r"C:\Users\admin\Dropbox\Research\개인연구\23_Brachy_BladderTocixity\GU toxicity_DB_python.xlsx",engine='openpyxl')
#df = pd.read_excel(r"C:\Users\admin\Dropbox\Research\개인연구\23_Brachy_BladderTocixity\GU toxicity_DB_python_norm-F3.xlsx",engine='openpyxl')
df.head()
df = df.sample(frac=1, random_state=1).reset_index(drop=True) # shffle and index reset
print(df)

input_original = df.iloc[:, 1:-1]
output_original = df.iloc[:, -1]
scaler = StandardScaler()
input_original = scaler.fit_transform(input_original)
patientID = df.iloc[:, 0]
patientID = patientID.to_numpy()
#input_original = input_original.to_numpy()

for iter1 in range(1, 6):
    print("Fold {} is under the training".format(iter1))
    kFoldParameters = 5
    currentFold = iter1
    output_original_length = output_original.__len__()
    rn_output_original = range(0, output_original_length)
    kf5 = KFold(n_splits=kFoldParameters, shuffle=False) # 5 is default

    training_indexes = {}
    test_indexes = {}
    counter = 1
    for train_index, test_index in kf5.split(rn_output_original):
        training_indexes["trainIndex_CV{0}".format(counter)] = train_index
        test_indexes["testIndex_CV{0}".format(counter)] = test_index
        counter = counter + 1
    # print("trainIndex:{}".format(train_index))
    # print("testIndex:{}".format(test_index))

    trainingDicKey_list = list(training_indexes.keys())
    trainingIndex_CV_F = training_indexes.get(trainingDicKey_list[currentFold - 1])

    testDicKey_list = list(test_indexes.keys())
    testIndex_CV_F = test_indexes.get(testDicKey_list[currentFold - 1])

    X_train = np.array(input_original[trainingIndex_CV_F])
    y_train = np.array(output_original[trainingIndex_CV_F])
    X_val = np.array(input_original[testIndex_CV_F])
    y_val = np.array(output_original[testIndex_CV_F])
    patientID_val = np.array(patientID[testIndex_CV_F])
    y_val_original = copy.deepcopy(y_val)

    print("Data is successfully loaded !!")
    print("Train input:{}, Train gt:{}".format(np.shape(X_train), np.shape(y_train)))
    print("Test input:{}, Test gt:{}".format(np.shape(X_val), np.shape(y_val)))

    # Binary classification
    y_train[y_train>0]=1
    y_val[y_val>0]=1
    # Class balance
    def get_class_distribution(obj):
        count_dict = {
            "normal": 0,
            "toxicity": 0,
        }

        for i in obj:
            if i == 0:
                count_dict['normal'] += 1
            elif i == 1:
                count_dict['toxicity'] += 1
            else:
                print("Check classes.")

        return count_dict

    fig, axes = plt.subplots(nrows=1, ncols=2, figsize=(16,7))
    # Train
    sns.barplot(data = pd.DataFrame.from_dict([get_class_distribution(y_train)]).melt(), x = "variable", y="value", hue="variable",  ax=axes[0]).set_title('Class Distribution in Train Set')
    # Validation
    sns.barplot(data = pd.DataFrame.from_dict([get_class_distribution(y_val)]).melt(), x = "variable", y="value", hue="variable",  ax=axes[1]).set_title('Class Distribution in Val Set')


    class ClassifierDataset(Dataset):

        def __init__(self, X_data, y_data):
            self.X_data = X_data
            self.y_data = y_data

        def __getitem__(self, index):
             x_data = self.X_data[index] + (torch.rand(self.X_data[index].size()[0]) * 0.2)
             return x_data, self.y_data[index]
            #return self.X_data[index], self.y_data[index]

        def __len__(self):
            return len(self.X_data)


    train_dataset = ClassifierDataset(torch.from_numpy(X_train).float(), torch.from_numpy(y_train).long())
    val_dataset = ClassifierDataset(torch.from_numpy(X_val).float(), torch.from_numpy(y_val).long())

    target_list = []
    for _, t in train_dataset:
        target_list.append(t)

    target_list = torch.tensor(target_list)
    target_list = target_list[torch.randperm(len(target_list))]

    class_count = [i for i in get_class_distribution(y_train).values()]
    class_weights = 1./torch.tensor(class_count, dtype=torch.float)
    class_weights_all = class_weights[target_list]

    weighted_sampler = WeightedRandomSampler(
        weights=class_weights_all,
        num_samples=len(class_weights_all),
        replacement=True
    )

    EPOCHS = 300
    BATCH_SIZE = 16
    LEARNING_RATE = 0.0001
    NUM_FEATURES = np.shape(input_original)[1]
    NUM_CLASSES = 2


    train_loader = DataLoader(dataset=train_dataset,
                              batch_size=BATCH_SIZE, sampler=weighted_sampler)
    val_loader = DataLoader(dataset=val_dataset, batch_size=1)


    class MulticlassClassification(nn.Module):
        def __init__(self, num_feature, num_class):
            super(MulticlassClassification, self).__init__()
            # #model 1
            # self.biasOp = False
            # self.layer_1 = nn.Linear(num_feature, 512, self.biasOp)
            # self.layer_2 = nn.Linear(512, 128, self.biasOp)
            # self.layer_3 = nn.Linear(128, 64, self.biasOp)
            # self.layer_out = nn.Linear(64, num_class, self.biasOp)
            #
            # self.relu = nn.ReLU()
            # self.LeakyReLU = nn.LeakyReLU()
            # self.dropout = nn.Dropout(p=0.2)
            # self.batchnorm1 = nn.BatchNorm1d(512)
            # self.batchnorm2 = nn.BatchNorm1d(128)
            # self.batchnorm3 = nn.BatchNorm1d(64)

            # # model 2
            # self.biasOp = False # False option is fixed.
            # self.layer_1 = nn.Linear(num_feature, 36, self.biasOp)
            # self.layer_2 = nn.Linear(36, 64, self.biasOp)
            # self.layer_3 = nn.Linear(64, 128, self.biasOp)
            # self.layer_4 = nn.Linear(128, 256, self.biasOp)
            # self.layer_5 = nn.Linear(256, 512, self.biasOp)
            # self.layer_6 = nn.Linear(512, 256, self.biasOp)
            # self.layer_7 = nn.Linear(256, 128, self.biasOp)
            # self.layer_8 = nn.Linear(128, 64, self.biasOp)
            # self.layer_9 = nn.Linear(64, 32, self.biasOp)
            # self.layer_out = nn.Linear(32, num_class, self.biasOp)
            #
            # self.relu = nn.ReLU()
            # self.LeakyReLU = nn.LeakyReLU()
            # self.dropout = nn.Dropout(p=0.2)
            # self.batchnorm1 = nn.BatchNorm1d(36)
            # self.batchnorm2 = nn.BatchNorm1d(64)
            # self.batchnorm3 = nn.BatchNorm1d(128)
            # self.batchnorm4 = nn.BatchNorm1d(256)
            # self.batchnorm5 = nn.BatchNorm1d(512)
            # self.batchnorm6 = nn.BatchNorm1d(256)
            # self.batchnorm7 = nn.BatchNorm1d(128)
            # self.batchnorm8 = nn.BatchNorm1d(64)
            # self.batchnorm9 = nn.BatchNorm1d(32)

            #model 3
            self.biasOp = False
            self.layer_1 = nn.Linear(num_feature, 36, self.biasOp)
            self.layer_2 = nn.Linear(36, 72, self.biasOp)
            self.layer_3 = nn.Linear(72, 144, self.biasOp)
            self.layer_4 = nn.Linear(144, 72, self.biasOp)
            self.layer_5 = nn.Linear(72, 36, self.biasOp)
            self.layer_out = nn.Linear(36, num_class, self.biasOp)

            self.relu = nn.ReLU()
            self.LeakyReLU = nn.LeakyReLU()
            self.dropout = nn.Dropout(p=0.2)
            self.batchnorm1 = nn.BatchNorm1d(36)
            self.batchnorm2 = nn.BatchNorm1d(72)
            self.batchnorm3 = nn.BatchNorm1d(144)
            self.batchnorm4 = nn.BatchNorm1d(72)
            self.batchnorm5 = nn.BatchNorm1d(36)

            # torch.nn.init.xavier_uniform_(self.layer_1.weight)
            # torch.nn.init.xavier_uniform_(self.layer_2.weight)
            # torch.nn.init.xavier_uniform_(self.layer_3.weight)
            # torch.nn.init.xavier_uniform_(self.layer_4.weight)
            # torch.nn.init.xavier_uniform_(self.layer_5.weight)


        def forward(self, x):
            # # model 1
            # x = self.layer_1(x)
            # x = self.batchnorm1(x)
            # x = self.LeakyReLU(x)
            #
            # x = self.layer_2(x)
            # x = self.batchnorm2(x)
            # x = self.LeakyReLU(x)
            # x = self.dropout(x)
            #
            # x = self.layer_3(x)
            # x = self.batchnorm3(x)
            # x = self.LeakyReLU(x)
            # x = self.dropout(x)
            #
            # x = self.layer_out(x)

            # # # model 2
            # x = self.layer_1(x)
            # x = self.batchnorm1(x)
            # x = self.LeakyReLU(x)
            #
            # x = self.layer_2(x)
            # x = self.batchnorm2(x)
            # x = self.LeakyReLU(x)
            # x = self.dropout(x)
            #
            # x = self.layer_3(x)
            # x = self.batchnorm3(x)
            # x = self.LeakyReLU(x)
            # x = self.dropout(x)
            #
            # x = self.layer_4(x)
            # x = self.batchnorm4(x)
            # x = self.LeakyReLU(x)
            # x = self.dropout(x)
            #
            # x = self.layer_5(x)
            # x = self.batchnorm5(x)
            # x = self.LeakyReLU(x)
            # x = self.dropout(x)
            #
            # x = self.layer_6(x)
            # x = self.batchnorm6(x)
            # x = self.LeakyReLU(x)
            # x = self.dropout(x)
            #
            # x = self.layer_7(x)
            # x = self.batchnorm7(x)
            # x = self.LeakyReLU(x)
            # x = self.dropout(x)
            #
            # x = self.layer_8(x)
            # x = self.batchnorm8(x)
            # x = self.LeakyReLU(x)
            # x = self.dropout(x)
            #
            # x = self.layer_9(x)
            # x = self.batchnorm9(x)
            # x = self.LeakyReLU(x)
            # x = self.dropout(x)
            #
            # x = self.layer_out(x)

            # model 1
            x = self.layer_1(x)
            x = self.batchnorm1(x)
            x = self.LeakyReLU(x)

            x = self.layer_2(x)
            x = self.batchnorm2(x)
            x = self.LeakyReLU(x)
            x = self.dropout(x)

            x = self.layer_3(x)
            x = self.batchnorm3(x)
            x = self.LeakyReLU(x)
            x = self.dropout(x)

            x = self.layer_4(x)
            x = self.batchnorm4(x)
            x = self.LeakyReLU(x)
            x = self.dropout(x)

            x = self.layer_5(x)
            x = self.batchnorm5(x)
            x = self.LeakyReLU(x)
            x = self.dropout(x)

            x = self.layer_out(x)


            return x



    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    print(device)



    model = MulticlassClassification(num_feature = NUM_FEATURES, num_class=NUM_CLASSES)
    model.to(device)

    criterion = nn.CrossEntropyLoss(weight=class_weights.to(device))
    optimizer = optim.Adam(model.parameters(), lr=LEARNING_RATE)
    print(model)


    def multi_acc(y_pred, y_test):
        y_pred_softmax = torch.log_softmax(y_pred, dim=1)
        _, y_pred_tags = torch.max(y_pred_softmax, dim=1)

        correct_pred = (y_pred_tags == y_test).float()
        acc = correct_pred.sum() / len(correct_pred)

        acc = torch.round(acc * 100)

        return acc


    accuracy_stats = {
        'train': [],
        "val": []
    }
    loss_stats = {
        'train': [],
        "val": []
    }


    predSetF =[]
    gtSetF = []
    bestScore  =0

    print("Begin training.")
    for e in tqdm(range(1, EPOCHS + 1)):
        # TRAINING
        train_epoch_loss = 0
        train_epoch_acc = 0
        model.train()
        train_len_counter = 0
        for X_train_batch, y_train_batch in train_loader:
            X_train_batch, y_train_batch = X_train_batch.to(device), y_train_batch.to(device)
            optimizer.zero_grad()

            y_train_pred = model(X_train_batch)

            train_loss = criterion(y_train_pred, y_train_batch)
            train_acc = multi_acc(y_train_pred, y_train_batch)

            train_loss.backward()
            optimizer.step()

            train_epoch_loss += train_loss.item()
            train_epoch_acc += train_acc.item()
            train_len_counter += 1

    # VALIDATION
            with torch.no_grad():
                val_epoch_loss = 0
                val_epoch_acc = 0
                predSet = []
                gtSet = []
                val_len_coutner = 0
                model.eval()
                for X_val_batch, y_val_batch in val_loader:
                    X_val_batch, y_val_batch = X_val_batch.to(device), y_val_batch.to(device)

                    y_val_pred = model(X_val_batch)

                    val_loss = criterion(y_val_pred, y_val_batch)
                    val_acc = multi_acc(y_val_pred, y_val_batch)

                    val_epoch_loss += val_loss.item()
                    val_epoch_acc += val_acc.item()
                    val_len_coutner += 1

                    # save the predicted and gt value
                    y_pred_softmax = torch.log_softmax(y_val_pred, dim=1)
                    _, y_pred_tags = torch.max(y_pred_softmax, dim=1)
                    y_val_pred_np = y_pred_tags.cpu().detach().numpy()
                    y_val_batch_np = y_val_batch.cpu().detach().numpy()
                    predSet.append(y_val_pred_np)
                    gtSet.append(y_val_batch_np)

        loss_stats['train'].append(train_epoch_loss / len(train_loader))
        loss_stats['val'].append(val_epoch_loss / len(val_loader))
        accuracy_stats['train'].append(train_epoch_acc / len(train_loader))
        accuracy_stats['val'].append(val_epoch_acc / len(val_loader))
        len_train = train_len_counter
        len_val = val_len_coutner
        currentScore = val_epoch_acc / len_val
        # print(bestScore)
        if bestScore <= currentScore:
            bestScore = currentScore
            predSetF = predSet
            gtSetF = gtSet
            print('Best score is updated !! ')

        print(f'Fold {iter1} | Epoch {e + 0:03}: | Train Loss: {train_epoch_loss / len(train_loader):.5f} | Val Loss: {val_epoch_loss / len(val_loader):.5f} | Train Acc: {train_epoch_acc / len(train_loader):.3f}| Val Acc: {val_epoch_acc / len(val_loader):.3f}')

    # plt.figure()
    # plt.plot(predSetF, 'r')
    # plt.plot(gtSetF)
    # plt.title('bestScore: {}'.format(bestScore))
    #
    #
    # # Create dataframes
    # train_val_acc_df = pd.DataFrame.from_dict(accuracy_stats).reset_index().melt(id_vars=['index']).rename(columns={"index":"epochs"})
    # train_val_loss_df = pd.DataFrame.from_dict(loss_stats).reset_index().melt(id_vars=['index']).rename(columns={"index":"epochs"})
    # # Plot the dataframes
    # fig, axes = plt.subplots(nrows=1, ncols=2, figsize=(20,7))
    # sns.lineplot(data=train_val_acc_df, x = "epochs", y="value", hue="variable",  ax=axes[0]).set_title('Train-Val Accuracy/Epoch')
    # sns.lineplot(data=train_val_loss_df, x = "epochs", y="value", hue="variable", ax=axes[1]).set_title('Train-Val Loss/Epoch')
    # plt.show()


    #ws = wb.active
    ws_write = wb.create_sheet(title="fold-{}".format(iter1))
    predSetF= list(predSetF)
    gtSetF = list(gtSetF)
    y_val_original = list(y_val_original)
    patientID_val = list(patientID_val)
    for i in range(len(predSet)):
        ws_write.cell(row=i+1, column=1).value = patientID_val[i]
        ws_write.cell(row=i+1, column=2).value = predSetF[i][0]
        ws_write.cell(row=i+1, column=3).value = gtSetF[i][0]
        ws_write.cell(row=i+1, column=4).value = y_val_original[i]

        #ws_write.append([predSet[i][0]])
        #ws_write.append([gtSet[i][1]])

saveFileName = os.path.join(r"ResultsReports", currentDT.strftime("%Y-%m-%d-%H-%M-%S-Results.xlsx"))
wb.save(filename=saveFileName)






