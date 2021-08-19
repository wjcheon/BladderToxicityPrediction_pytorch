import numpy as np
import pandas as pd
from tqdm.notebook import tqdm
import matplotlib.pyplot as plt
import seaborn as sns

import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader, WeightedRandomSampler

from sklearn.preprocessing import MinMaxScaler, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix, classification_report
from sklearn.model_selection import KFold
import openpyxl

df = pd.read_excel(r"C:\Users\admin\Dropbox\Research\개인연구\23_Brachy_BladderTocixity\GU toxicity_DB_python-balance.xlsx",engine='openpyxl')
#df = pd.read_excel(r"C:\Users\admin\Dropbox\Research\개인연구\23_Brachy_BladderTocixity\GU toxicity_DB_python.xlsx",engine='openpyxl')
df.head()


input_original = df.iloc[:, 1:-1]
output_original = df.iloc[:, -1]
scaler = StandardScaler()
input_original = scaler.fit_transform(input_original)

kFoldParameters = 5
currentFold =  1
output_original_length = output_original.__len__()
rn_output_original = range(0, output_original_length)
kf5 = KFold(n_splits=kFoldParameters, shuffle=False) # 5 is default

training_indexes = {}
test_indexes = {}
counter = 1
for train_index, test_index in kf5.split(rn_output_original):
    training_indexes["trainIndex_CV{0}".format(counter)] = train_index
    test_indexes["testIndex_CV{0}".format(counter)] = test_index
    counter = counter + 1
# print("trainIndex:{}".format(train_index))
# print("testIndex:{}".format(test_index))

trainingDicKey_list = list(training_indexes.keys())
trainingIndex_CV_F = training_indexes.get(trainingDicKey_list[currentFold - 1])

testDicKey_list = list(test_indexes.keys())
testIndex_CV_F = test_indexes.get(testDicKey_list[currentFold - 1])

X_train = np.array(input_original[trainingIndex_CV_F])
y_train = np.array(output_original[trainingIndex_CV_F])
X_val = np.array(input_original[testIndex_CV_F])
y_val = np.array(output_original[testIndex_CV_F])

print("Data is successfully loaded !!")
print("Train input:{}, Train gt:{}".format(np.shape(X_train), np.shape(y_train)))
print("Test input:{}, Test gt:{}".format(np.shape(X_val), np.shape(y_val)))


#
# X_trainval, X_test, y_trainval, y_test = train_test_split(X, y, test_size=0.1, random_state=69)
#
# scaler = MinMaxScaler()
# X_train = scaler.fit_transform(X_trainval)
# X_val = scaler.transform(X_test)
# X_train, y_train = np.array(X_train), np.array(y_trainval)
# X_val, y_val = np.array(X_val), np.array(y_test)


class ClassifierDataset(Dataset):

    def __init__(self, X_data, y_data):
        self.X_data = X_data
        self.y_data = y_data

    def __getitem__(self, index):

        # x_data = self.X_data[index] + (torch.rand(self.X_data[index].size()[0]) * 0.01)
        # return x_data, self.y_data[index]
        return self.X_data[index], self.y_data[index]

    def __len__(self):
        return len(self.X_data)


#train_dataset = ClassifierDataset(torch.from_numpy(X_train).float(), torch.from_numpy(y_train).long())
#val_dataset = ClassifierDataset(torch.from_numpy(X_val).float(), torch.from_numpy(y_val).long())
train_dataset = ClassifierDataset(torch.from_numpy(X_train).float(), torch.from_numpy(y_train).float())
val_dataset = ClassifierDataset(torch.from_numpy(X_val).float(), torch.from_numpy(y_val).float())

target_list = []
for _, t in train_dataset:
    target_list.append(t)

target_list = torch.tensor(target_list)
target_list = target_list[torch.randperm(len(target_list))]




EPOCHS = 300
BATCH_SIZE = 16
LEARNING_RATE = 0.0007
beta1 = 0.5
beta2 = 0.999
num_epochs_decay = 200
NUM_FEATURES = np.shape(input_original)[1]
NUM_CLASSES = 1


train_loader = DataLoader(dataset=train_dataset,batch_size=BATCH_SIZE, shuffle=True)
val_loader = DataLoader(dataset=val_dataset, batch_size=1, shuffle=False)


class MulticlassClassification(nn.Module):
    def __init__(self, num_feature, num_class):
        super(MulticlassClassification, self).__init__()
        self.biasOp = False
        self.layer_1 = nn.Linear(num_feature, 512, bias=self.biasOp)
        self.layer_2 = nn.Linear(512, 128, bias=self.biasOp)
        self.layer_3 = nn.Linear(128, 64, bias=self.biasOp)
        self.layer_out = nn.Linear(64, num_class, bias=self.biasOp)

        self.relu = nn.ReLU()
        self.dropout = nn.Dropout(p=0.2)
        self.batchnorm1 = nn.BatchNorm1d(512)
        self.batchnorm2 = nn.BatchNorm1d(128)
        self.batchnorm3 = nn.BatchNorm1d(64)

    def forward(self, x):
        x = self.layer_1(x)
        x = self.batchnorm1(x)
        x = self.relu(x)

        x = self.layer_2(x)
        x = self.batchnorm2(x)
        x = self.relu(x)
        x = self.dropout(x)

        x = self.layer_3(x)
        x = self.batchnorm3(x)
        x = self.relu(x)
        x = self.dropout(x)

        x = self.layer_out(x)

        return x



device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
print(device)

model = MulticlassClassification(num_feature = NUM_FEATURES, num_class=NUM_CLASSES)
model.to(device)

criterion = nn.CrossEntropyLoss()
criterion = nn.MSELoss()
#optimizer = optim.Adam(list(model.parameters()), LEARNING_RATE, [beta1, beta2])
optimizer = optim.Adam(model.parameters())
print(model)


def multi_acc(y_pred, y_test):
    y_pred_softmax = torch.log_softmax(y_pred, dim=1)
    _, y_pred_tags = torch.max(y_pred_softmax, dim=1)

    correct_pred = (y_pred_tags == y_test).float()
    acc = correct_pred.sum() / len(correct_pred)

    acc = torch.round(acc * 100)

    return acc


accuracy_stats = {
    'train': [],
    "val": []
}
loss_stats = {
    'train': [],
    "val": []
}




print("Begin training.")
for e in tqdm(range(1, EPOCHS + 1)):
    # TRAINING
    train_epoch_loss = 0
    train_epoch_acc = 0
    model.train()
    for X_train_batch, y_train_batch in train_loader:
        X_train_batch, y_train_batch = X_train_batch.to(device), y_train_batch.to(device)
        optimizer.zero_grad()

        y_train_pred = model(X_train_batch)

        train_loss = criterion(y_train_pred, y_train_batch)
        #train_acc = multi_acc(y_train_pred, y_train_batch)

        train_loss.backward()
        optimizer.step()

        train_epoch_loss += train_loss.item()
        #train_epoch_acc += train_acc.item()

        # if (e + 1) > (EPOCHS - num_epochs_decay):
        #     LEARNING_RATE -= (LEARNING_RATE / float(num_epochs_decay))
        #     for param_group in optimizer.param_groups:
        #         param_group['lr'] = LEARNING_RATE
        #     print('Decay learning rate to lr: {}.'.format(LEARNING_RATE))

# VALIDATION
        with torch.no_grad():
            val_epoch_loss = 0
            val_epoch_acc = 0
            predSet = []
            gtSet = []
            model.eval()
            for X_val_batch, y_val_batch in val_loader:
                X_val_batch, y_val_batch = X_val_batch.to(device), y_val_batch.to(device)

                y_val_pred = model(X_val_batch)

                val_loss = criterion(y_val_pred, y_val_batch)
                #val_acc = multi_acc(y_val_pred, y_val_batch)

                val_epoch_loss += val_loss.item()
                #val_epoch_acc += val_acc.item()

                # save the predicted and gt value
                y_val_pred_np = y_val_pred.cpu().detach().numpy()
                y_val_batch_np = y_val_batch.cpu().detach().numpy()
                predSet.append(y_val_pred_np)
                gtSet.append(y_val_batch_np)

    loss_stats['train'].append(train_epoch_loss / len(train_loader))
    loss_stats['val'].append(val_epoch_loss / len(val_loader))
    #accuracy_stats['train'].append(train_epoch_acc / len(train_loader))
    #accuracy_stats['val'].append(val_epoch_acc / len(val_loader))

    #print(f'Epoch {e + 0:03}: | Train Loss: {train_epoch_loss / len(train_loader):.5f} | Val Loss: {val_epoch_loss / len(val_loader):.5f} | Train Acc: {train_epoch_acc / len(train_loader):.3f}| Val Acc: {val_epoch_acc / len(val_loader):.3f}')
    print(f'Epoch {e + 0:03}: | Train Loss: {train_epoch_loss / len(train_loader):.5f} | Val Loss: {val_epoch_loss / len(val_loader):.5f}')

wb = openpyxl.Workbook()
ws = wb.active
#ws_write = wb.create_sheet(0)
predSet= list(predSet)
gtSet = list(gtSet)
for i in range(len(predSet)):
    ws.cell(row=i+1, column=1).value = predSet[i][0][0]
    ws.cell(row=i+1, column=2).value = gtSet[i][0]
    #ws_write.append([predSet[i][0]])
    #ws_write.append([gtSet[i][1]])
wb.save(filename='data.xlsx')

# Create dataframes
train_val_acc_df = pd.DataFrame.from_dict(accuracy_stats).reset_index().melt(id_vars=['index']).rename(columns={"index":"epochs"})
train_val_loss_df = pd.DataFrame.from_dict(loss_stats).reset_index().melt(id_vars=['index']).rename(columns={"index":"epochs"})
# Plot the dataframes
fig, axes = plt.subplots(nrows=1, ncols=2, figsize=(20,7))
sns.lineplot(data=train_val_acc_df, x = "epochs", y="value", hue="variable",  ax=axes[0]).set_title('Train-Val Accuracy/Epoch')
sns.lineplot(data=train_val_loss_df, x = "epochs", y="value", hue="variable", ax=axes[1]).set_title('Train-Val Loss/Epoch')
plt.show()





